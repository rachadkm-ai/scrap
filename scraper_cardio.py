"""
Scraper pour etablissements.fhf.fr
Basé sur l'analyse réelle du HTML de la page.

Page 1 : #Groupe_1811 contient des <a href="javascript:void(0);" data-id="XX">
         Chaque <a> a data-id (code département) et data-bs-original-title (nom)
         On construit l'URL : /annuaire/search?type=etablissement&department=XX

Page 2 : .card-body.d-flex.flex-column.justify-content-between
         → h3.card-title = nom établissement
         → a.btn-link href="/annuaire/member/..." = lien fiche

Page 3 : Accordéon "Services de l'établissement"
         → .row-item → .title + .col h3 + .col a[href^="mailto:"]

Installation :
    pip install playwright openpyxl
    python -m playwright install chromium

Lancement :
    python scraper_fhf_v2.py
"""

import asyncio
import csv
import json
from pathlib import Path

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

BASE_URL   = "https://etablissements.fhf.fr"
OUTPUT_DIR = Path(".")
HEADLESS   = True
DELAY      = 1.5   # secondes entre requetes (respecter le serveur)


async def get_group_data_ids(page) -> list:
    """
    Retourne la liste complète des départements français de manière statique.
    Évite les soucis de sélecteurs introuvables et de temps d'attente.
    """
    deps = {
        "01": "Ain", "02": "Aisne", "03": "Allier", "04": "Alpes-de-Haute-Provence",
        "05": "Hautes-Alpes", "06": "Alpes-Maritimes", "07": "Ardèche", "08": "Ardennes",
        "09": "Ariège", "10": "Aube", "11": "Aude", "12": "Aveyron",
        "13": "Bouches-du-Rhône", "14": "Calvados", "15": "Cantal", "16": "Charente",
        "17": "Charente-Maritime", "18": "Cher", "19": "Corrèze", "2A": "Corse-du-Sud",
        "2B": "Haute-Corse", "21": "Côte-d'Or", "22": "Côtes-d'Armor", "23": "Creuse",
        "24": "Dordogne", "25": "Doubs", "26": "Drôme", "27": "Eure", "28": "Eure-et-Loir",
        "29": "Finistère", "30": "Gard", "31": "Haute-Garonne", "32": "Gers",
        "33": "Gironde", "34": "Hérault", "35": "Ille-et-Vilaine", "36": "Indre",
        "37": "Indre-et-Loire", "38": "Isère", "39": "Jura", "40": "Landes",
        "41": "Loir-et-Cher", "42": "Loire", "43": "Haute-Loire", "44": "Loire-Atlantique",
        "45": "Loiret", "46": "Lot", "47": "Lot-et-Garonne", "48": "Lozère",
        "49": "Maine-et-Loire", "50": "Manche", "51": "Marne", "52": "Haute-Marne",
        "53": "Mayenne", "54": "Meurthe-et-Moselle", "55": "Meuse", "56": "Morbihan",
        "57": "Moselle", "58": "Nièvre", "59": "Nord", "60": "Oise", "61": "Orne",
        "62": "Pas-de-Calais", "63": "Puy-de-Dôme", "64": "Pyrénées-Atlantiques",
        "65": "Hautes-Pyrénées", "66": "Pyrénées-Orientales", "67": "Bas-Rhin",
        "68": "Haut-Rhin", "69": "Rhône", "70": "Haute-Saône", "71": "Saône-et-Loire",
        "72": "Sarthe", "73": "Savoie", "74": "Haute-Savoie", "75": "Paris",
        "76": "Seine-Maritime", "77": "Seine-et-Marne", "78": "Yvelines", "79": "Deux-Sèvres",
        "80": "Somme", "81": "Tarn", "82": "Tarn-et-Garonne", "83": "Var", "84": "Vaucluse",
        "85": "Vendée", "86": "Vienne", "87": "Haute-Vienne", "88": "Vosges", "89": "Yonne",
        "90": "Territoire de Belfort", "91": "Essonne", "92": "Hauts-de-Seine",
        "93": "Seine-Saint-Denis", "94": "Val-de-Marne", "95": "Val-d'Oise",
        "971": "Guadeloupe", "972": "Martinique", "973": "Guyane", "974": "La Réunion",
        "976": "Mayotte", "987": "Polynésie Française", "988": "Nouvelle-Calédonie"
    }
    
    groups = [{"data_id": k, "title": v} for k, v in deps.items()]
    print(f"[Etape 1] {len(groups)} departements charges statiquement.")
    return groups


# ──────────────────────────────────────────────
# ÉTAPE 2 : Page de liste par departement (avec pagination)
# ──────────────────────────────────────────────
async def get_cards_for_group(page, group: dict) -> list:
    """
    Navigue vers la page de recherche par departement.
    Extrait toutes les cartes avec gestion de la pagination.
    """
    dept_id    = group["data_id"]
    base_url   = f"{BASE_URL}/annuaire/search?type=etablissement&department={dept_id}"
    cards_all  = []
    page_num   = 0

    while True:
        url = f"{base_url}&page={page_num}" if page_num > 0 else base_url
        try:
            await page.goto(url, wait_until="domcontentloaded")
        except Exception as e:
            print(f"    Erreur navigation {url}: {e}")
            break

        card_els = await page.query_selector_all(
            ".card-body.d-flex.flex-column.justify-content-between"
        )
        if not card_els:
            break

        for el in card_els:
            h3_el   = await el.query_selector("h3.card-title")
            link_el = await el.query_selector("a.btn-link")
            h3_text = (await h3_el.inner_text()).strip() if h3_el   else ""
            href    = await link_el.get_attribute("href")  if link_el else ""
            if href:
                if not href.startswith("http"):
                    href = BASE_URL + href
                cards_all.append({"h3": h3_text, "card_link": href})

        # Verifier s'il y a une page suivante
        next_btn = await page.query_selector(
            "a[rel='next'], .pager__item--next a, li.next a, a.page-link[aria-label='Next']"
        )
        if not next_btn:
            break
        page_num += 1
        await asyncio.sleep(DELAY)

    print(f"  [Etape 2] {len(cards_all)} cartes pour {group['title']} (id={dept_id})")
    return cards_all


# ──────────────────────────────────────────────
# ÉTAPE 3 : Fiche etablissement - services
# ──────────────────────────────────────────────
async def get_services(page, detail_url: str) -> list:
    """
    Navigue vers la fiche etablissement.
    Ouvre l'accordeon 'Services de l'etablissement' et extrait les services de cardiologie.
    Analyse la description du service pour trouver les mots-clés liés au plateau technique.
    """
    try:
        await page.goto(detail_url, wait_until="domcontentloaded")
    except Exception as e:
        print(f"    Erreur {detail_url}: {e}")
        return []

    # Chercher et cliquer le bouton accordeon "Services de l'etablissement"
    try:
        btns = await page.query_selector_all("button.accordion-button")
        for btn in btns:
            text = (await btn.inner_text()).strip()
            if "services" in text.lower() or "direction" in text.lower():
                css_class = await btn.get_attribute("class") or ""
                if "collapsed" in css_class:
                    await btn.click()
                    await page.wait_for_timeout(900)
    except Exception:
        pass

    # Extraire les .row-item et filtrer pour la cardiologie
    row_els  = await page.query_selector_all(".row-item")
    services = []
    email_am = ""

    for el in row_els:
        title_el    = await el.query_selector(".title")
        h3_el       = await el.query_selector(".col h3")
        mailto_el   = await el.query_selector('.col a[href^="mailto:"]')

        title       = (await title_el.inner_text()).strip()    if title_el    else ""
        svc_h3      = (await h3_el.inner_text()).strip()       if h3_el       else ""
        mailto_href = await mailto_el.get_attribute("href")    if mailto_el   else ""
        mailto      = mailto_href.replace("mailto:", "")       if mailto_href else ""

        t_low = title.lower()
        h3_low = svc_h3.lower()
        
        # Détection "Direction des Affaires Médicales"
        # Un hôpital n'en a qu'un seul, donc on ne stocke que le premier email trouvé
        is_am = ("affaires" in t_low and "dical" in t_low) or ("affaires" in h3_low and "dical" in h3_low)
        if is_am and mailto and not email_am:
            email_am = mailto

        # Filtre Cardiologie
        if "cardio" not in t_low and "cardio" not in h3_low:
            continue

        # Recherche de plateau technique dans TOUT le texte du service (incluant la description)
        full_text = (await el.inner_text()).lower()
        has_interv = "coronarographie" in full_text or "cardiologie interventionnelle" in full_text
        plateau_info = "Oui" if has_interv else "Non spécifié / Non"

        services.append({
            "title": title, 
            "service_h3": svc_h3, 
            "mailto": mailto,
            "plateau_cardio": plateau_info
        })

    return {"services": services, "email_am": email_am}


# ──────────────────────────────────────────────
# ORCHESTRATION PRINCIPALE
# ──────────────────────────────────────────────
async def scrape() -> list:
    results = []

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        )
        page = await context.new_page()

        # Etape 1
        groups = await get_group_data_ids(page)

        for g_idx, group in enumerate(groups, 1):
            print(f"\n[Groupe {g_idx}/{len(groups)}] {group['title']} (id={group['data_id']})")
            await asyncio.sleep(DELAY)

            # Etape 2
            cards = await get_cards_for_group(page, group)

            for c_idx, card in enumerate(cards, 1):
                if not card["card_link"]:
                    continue
                print(f"  [Carte {c_idx}/{len(cards)}] {card['h3']}")
                await asyncio.sleep(DELAY)

                # Etape 3
                svc_data = await get_services(page, card["card_link"])
                services = svc_data.get("services", [])
                email_am = svc_data.get("email_am", "")

                if services:
                    for svc in services:
                        results.append({
                            "groupe_id":     group["data_id"],
                            "groupe_titre":  group["title"],
                            "etablissement": card["h3"],
                            "fiche_url":     card["card_link"],
                            "service_titre": svc["title"],
                            "service_h3":    svc["service_h3"],
                            "email":         svc["mailto"],
                            "email_affaires": email_am,
                            "plateau_cardio": svc["plateau_cardio"]
                        })
                # On retire le "else" ici car si un établissement n'a pas de cardiologie,
                # on ne veut pas l'ajouter du tout au fichier de résultats.

        await browser.close()

    print(f"\nScraping termine : {len(results)} lignes collectees.")
    return results


# ──────────────────────────────────────────────
# EXPORTS CSV / JSON / EXCEL
# ──────────────────────────────────────────────
def export_csv(data, path):
    if not data:
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    print(f"CSV   sauvegarde : {path}")


def export_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON  sauvegarde : {path}")


def export_excel(data, path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl non installe. Lancez : pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FHF Etablissements"

    if not data:
        wb.save(path)
        return

    headers = list(data[0].keys())
    labels  = {
        "groupe_id":     "ID Groupe",
        "groupe_titre":  "Groupe / Departement",
        "etablissement": "Etablissement",
        "fiche_url":     "URL Fiche",
        "service_titre": "Titre Service",
        "service_h3":    "Nom Service",
        "email":         "Email Service",
        "email_affaires": "Email Affaires Médicales",
        "plateau_cardio":"Plateau Coronaro/Interventionnel",
    }

    fill   = PatternFill("solid", fgColor="2E75B6")
    font_h = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=labels.get(h, h))
        cell.fill      = fill
        cell.font      = font_h
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(data, 2):
        for ci, h in enumerate(headers, 1):
            val  = row.get(h, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            if h in ["email", "email_affaires"] and val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(color="0563C1", underline="single")
            elif h == "fiche_url" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    widths = [10, 28, 42, 48, 25, 32, 38, 38, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
    print(f"Excel sauvegarde : {path}")


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
if __name__ == "__main__":
    data = asyncio.run(scrape())
    if data:
        export_csv(data,   OUTPUT_DIR / "fhf_cardiologie.csv")
        export_json(data,  OUTPUT_DIR / "fhf_cardiologie.json")
        export_excel(data, OUTPUT_DIR / "fhf_cardiologie.xlsx")
    else:
        print("Aucune donnee collectee.")
